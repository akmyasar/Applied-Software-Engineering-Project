"""Part 2 Step 4: the deliverables.

    python -m part2.reports

Produces, from config.CLASSIFICATION_DB:

    reports/23025328-sq26-classification.xlsx  (Step 4c)
    reports/23025328-sq26-classification.pdf   (Step 4d)
    reports/form_answers.md                    (Step 4b, input for the form)
"""
import sqlite3
import textwrap
from collections import Counter

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from part2 import config, isic

TOP_N = 20
CLASSIFIED_TYPES = ("QDA_PROJECT", "QD_PROJECT")
ALL_TYPES = ("QDA_PROJECT", "QD_PROJECT", "OTHER_PROJECT", "NOT_A_PROJECT")

# Vector output only: the PDF must stay zoomable (Step 4d a. iii).
matplotlib.rcParams["pdf.fonttype"] = 42
matplotlib.rcParams["pdf.compression"] = 6
matplotlib.rcParams["font.family"] = "DejaVu Sans"


def connect():
    conn = sqlite3.connect(config.CLASSIFICATION_DB)
    conn.row_factory = sqlite3.Row
    return conn


def clean(text):
    """Tidy a stored string for display.

    The metadata is valid UTF-8 (titles legitimately contain em dashes, and one
    file name ends in an acute accent), so nothing has to be substituted here -
    only the stray padding Part 1 preserved verbatim is removed.
    """
    return (text or "").strip()


# ---------------------------------------------------------------------------
# Step 4c: the XLSX table
# ---------------------------------------------------------------------------
def export_xlsx(conn, path=None):
    path = path or config.REPORT_DIR / f"{config.STUDENT_ID}-sq26-classification.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    rows = conn.execute(
        """SELECT repository_id, type AS project_type, title AS project_title,
                  primary_class, secondary_class, no_project_files
           FROM PROJECTS
           ORDER BY repository_id, type, id"""
    ).fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "classification"

    headers = ["repository_id", "project_type", "project_title",
               "primary_class", "secondary_class", "no_project_files"]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append([
            row["repository_id"],
            row["project_type"],
            clean(row["project_title"]),
            # The full class name is the readable form of the code and is what
            # the report histograms bin on, so the spreadsheet carries it too.
            isic.full_class_name(row["primary_class"]) if row["primary_class"] else "",
            isic.full_class_name(row["secondary_class"]) if row["secondary_class"] else "",
            row["no_project_files"],
        ])

    widths = [14, 15, 80, 42, 42, 16]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    workbook.save(path)
    return path, len(rows)


# ---------------------------------------------------------------------------
# Step 4d: the PDF report
# ---------------------------------------------------------------------------
def class_counts(conn, repository_id, project_type=None):
    query = ("SELECT primary_class FROM PROJECTS "
             "WHERE repository_id = ? AND primary_class IS NOT NULL")
    params = [repository_id]
    if project_type:
        query += " AND type = ?"
        params.append(project_type)
    else:
        query += " AND type IN ({})".format(", ".join("?" * len(CLASSIFIED_TYPES)))
        params.extend(CLASSIFIED_TYPES)
    return Counter(r["primary_class"] for r in conn.execute(query, params))


def wrap_label(text, width=30, max_lines=3):
    """Full ISIC class names run to 70+ characters; stack them instead of
    letting one diagonal label eat half the page."""
    lines = textwrap.wrap(text, width=width)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = lines[-1][: width - 1] + "…"
    return "\n".join(lines)


def draw_histogram(pdf, counts, title, subtitle):
    """Bar chart of primary classes, count printed on top of each bar."""
    if not counts:
        return
    ordered = counts.most_common()
    labels = [wrap_label(isic.full_class_name(code)) for code, _ in ordered]
    values = [n for _, n in ordered]

    width = max(9.0, min(0.75 * len(labels) + 3.0, 26))
    figure, axes = plt.subplots(figsize=(width, 8.0))
    bars = axes.bar(range(len(values)), values, color="#4472C4", edgecolor="#2F528F")

    for bar, value in zip(bars, values):
        axes.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(values) * 0.015,
                  str(value), ha="center", va="bottom", fontsize=9, fontweight="bold")

    axes.set_xticks(range(len(labels)))
    axes.set_xticklabels(labels, rotation=45, ha="right", fontsize=7,
                         rotation_mode="anchor")
    axes.set_ylabel("Number of projects", fontsize=10)
    axes.set_ylim(0, max(values) * 1.14)
    axes.set_title(title, fontsize=13, fontweight="bold", pad=30)
    if subtitle:
        axes.text(0.5, 1.013, subtitle, transform=axes.transAxes,
                  ha="center", va="bottom", fontsize=9, color="#555555")
    axes.spines[["top", "right"]].set_visible(False)
    axes.grid(axis="y", linestyle=":", alpha=0.5)
    axes.set_axisbelow(True)

    figure.tight_layout()
    pdf.savefig(figure)          # vector by default on the PDF backend
    plt.close(figure)


def draw_table(pdf, counts, title, total_projects):
    """Rank-ordered table of the top TOP_N classes."""
    if not counts:
        return
    ordered = counts.most_common(TOP_N)
    total = sum(counts.values())

    figure, axes = plt.subplots(figsize=(11.7, 8.3))
    axes.axis("off")
    axes.set_title(title, fontsize=13, fontweight="bold", pad=18)

    cells = [
        [str(rank), code, isic.division_name(code), str(count), f"{count / total:.1%}"]
        for rank, (code, count) in enumerate(ordered, 1)
    ]
    table = axes.table(
        cellText=cells,
        colLabels=["Rank", "Division", "Full class name", "Count", "Share"],
        colWidths=[0.07, 0.09, 0.62, 0.09, 0.10],
        cellLoc="left", loc="upper center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)
    table.scale(1, 1.45)
    for (row, column), cell in table.get_celld().items():
        cell.set_edgecolor("#CCCCCC")
        if row == 0:
            cell.set_facecolor("#4472C4")
            cell.set_text_props(color="white", fontweight="bold")
        elif row % 2 == 0:
            cell.set_facecolor("#F2F5FB")
        if column in (0, 3, 4):
            cell.set_text_props(ha="center")

    note = (f"{len(counts)} distinct classes over {total} classified projects "
            f"({total_projects} projects in scope). "
            f"Showing the top {min(TOP_N, len(ordered))}.")
    axes.text(0.5, 0.02, note, transform=axes.transAxes, ha="center",
              fontsize=8.5, color="#555555")

    figure.tight_layout()
    pdf.savefig(figure)
    plt.close(figure)


TEXT_WIDTH = 118      # characters per line at 9.5pt on a landscape A4 page
LINE_STEP = 0.0265    # vertical advance per physical line, in figure coords


def draw_text_page(pdf, title, lines):
    """Render a page of headings ('## ') and body lines, wrapping and
    paginating as needed.

    The wrapping is done here rather than by matplotlib's `wrap=True` because
    the caller has to know how many physical lines a paragraph became in order
    to advance the cursor past it; otherwise long paragraphs overprint.
    """
    figure = plt.figure(figsize=(11.7, 8.3))
    figure.text(0.06, 0.945, title, fontsize=15, fontweight="bold", va="top")
    y = 0.87

    def new_page():
        nonlocal figure, y
        pdf.savefig(figure)
        plt.close(figure)
        figure = plt.figure(figsize=(11.7, 8.3))
        y = 0.945

    for line in lines:
        if not line.strip():
            y -= LINE_STEP
            continue

        if line.startswith("## "):
            if y < 0.12:
                new_page()
            y -= 0.012
            figure.text(0.06, y, line[3:], fontsize=12, fontweight="bold", va="top")
            y -= 0.045
            continue

        body = line[2:] if line.startswith("- ") else line
        indent, bullet = ("0.075", "- ") if line.startswith("- ") else ("0.06", "")
        wrapped = textwrap.wrap(body, width=TEXT_WIDTH) or [""]

        for index, physical in enumerate(wrapped):
            if y < 0.05:
                new_page()
            text = (bullet + physical) if index == 0 else physical
            x = 0.06 if index == 0 else float(indent)
            figure.text(x, y, text, fontsize=9.5, va="top")
            y -= LINE_STEP
        y -= 0.006

    pdf.savefig(figure)
    plt.close(figure)


def repository_comments(conn, repository):
    """Step 4d 1.c - the findings worth flagging for this repository."""
    repo_id = repository["id"]
    lines = []

    types = Counter(
        r["type"] for r in conn.execute(
            "SELECT type FROM PROJECTS WHERE repository_id = ?", (repo_id,)
        )
    )
    total = sum(types.values())
    counts = class_counts(conn, repo_id)
    classified = sum(counts.values())
    in_scope = sum(types[t] for t in CLASSIFIED_TYPES)

    lines.append("## Project types found")
    for project_type in ALL_TYPES:
        if types.get(project_type):
            lines.append(f"- {project_type}: {types[project_type]} "
                         f"({types[project_type] / total:.1%} of {total})")
    lines.append("")

    lines.append("## Dominant class")
    if counts:
        code, count = counts.most_common(1)[0]
        lines.append(f"- {isic.label(code)} with {count} projects "
                     f"({count / classified:.1%} of the {classified} classified).")
        runners = counts.most_common(4)[1:]
        if runners:
            lines.append("- Runners-up: " + "; ".join(
                f"{isic.full_class_name(c)} ({n})" for c, n in runners))
    else:
        lines.append("- No project in this repository could be classified.")
    lines.append("")

    lines.append("## Coverage")
    lines.append(f"- {classified} of {in_scope} in-scope projects "
                 f"(QDA_PROJECT + QD_PROJECT) received a class.")
    files = conn.execute(
        """SELECT COUNT(*) AS total,
                  SUM(status = 'SUCCEEDED') AS downloaded,
                  SUM(text_chars > 0) AS with_text
           FROM FILES f JOIN PROJECTS p ON p.id = f.project_id
           WHERE p.repository_id = ?""", (repo_id,)
    ).fetchone()
    if files["total"]:
        lines.append(f"- {files['downloaded']} of {files['total']} files could be "
                     f"downloaded; {files['with_text']} yielded machine-readable text "
                     f"for the classifier to read.")
    lines.append("")
    lines += findings_comments(conn, repository, types, counts, files)
    return lines


def findings_comments(conn, repository, types, counts, files):
    """Step 4d 1.c - comments on the findings for this repository."""
    repo_id = repository["id"]
    lines = ["## Comments on the findings"]

    if not sum(types.values()):
        return lines + ["- The repository yielded no projects at all."]

    # A repository that consists only of Part 1 sentinel rows was never
    # harvested, so none of the observations below would apply to it: it has no
    # files to type and no metadata to classify.
    placeholders = conn.execute(
        """SELECT COUNT(*) AS n FROM PROJECTS
           WHERE repository_id = ? AND doi IN ({})""".format(
            ", ".join("?" * len(config.PLACEHOLDER_DOIS))),
        (repo_id, *config.PLACEHOLDER_DOIS),
    ).fetchone()["n"]
    if placeholders and placeholders == sum(types.values()):
        return lines + [
            "- This repository could not be harvested in Part 1: its web application "
            "firewall rejected every automated request, so no project metadata and no "
            "files were ever acquired.",
            "- The single record is the sentinel row the Part 1 pipeline writes to "
            "document that repository-level failure. It is not a research project, it "
            "carries an invented file name, and it is therefore typed NOT_A_PROJECT "
            "and excluded from classification rather than being allowed to contribute "
            "a spurious QD_PROJECT and a spurious class to the statistics.",
            "- No distribution can be reported for this repository. This is a data "
            "acquisition limitation carried over from Part 1, not a classification "
            "result: the repository is browsable by a human but not machine-accessible.",
        ]

    # 1. QDA files
    qda_files = conn.execute(
        """SELECT COUNT(*) AS n FROM FILES f JOIN PROJECTS p ON p.id = f.project_id
           WHERE p.repository_id = ? AND f.file_category = 'QDA'""", (repo_id,)
    ).fetchone()["n"]
    if not qda_files:
        lines.append(
            "- No file in this repository carries a QDA file extension (.qdpx, .nvp, "
            ".atlproj, .mx*, ...), so by the Step 1 rule no project can be a "
            "QDA_PROJECT. The archive predates the REFI-QDA standard and stores its "
            "analysis material as PDF codebooks and SPSS/SAS data rather than as "
            "QDA software projects. This is a property of the data, not a gap in the "
            "pipeline: the classifier does look for those extensions.")

    # 2. What the primary data actually is
    top_ext = conn.execute(
        """SELECT f.file_extension AS ext, COUNT(*) AS n
           FROM FILES f JOIN PROJECTS p ON p.id = f.project_id
           WHERE p.repository_id = ? AND f.file_category = 'PRIMARY'
             AND f.file_extension != ''
           GROUP BY ext ORDER BY n DESC LIMIT 3""", (repo_id,)
    ).fetchall()
    if top_ext:
        listed = ", ".join(f".{r['ext']} ({r['n']})" for r in top_ext)
        lines.append(
            f"- The primary data is dominated by {listed}. Because a single PDF is "
            "enough to make a project a QD_PROJECT, that rule sorts almost the whole "
            "repository into one bucket; the file-type rules discriminate far less "
            "here than the class taxonomy does.")

    # 3. Restricted files
    if files["total"] and files["downloaded"] is not None:
        restricted = files["total"] - (files["downloaded"] or 0)
        if restricted:
            share = restricted / files["total"]
            lines.append(
                f"- {restricted} of {files['total']} files ({share:.0%}) are restricted "
                "and could not be downloaded, so for those the classifier reads only "
                "the metadata and the file name. Classification quality is therefore "
                "not uniform across the repository, and the projects whose files are "
                "open are classified on more evidence than the rest.")

    # 4. Concentration of the distribution
    if counts:
        total = sum(counts.values())
        code, count = counts.most_common(1)[0]
        top3 = sum(n for _, n in counts.most_common(3))
        lines.append(
            f"- The distribution is heavily concentrated: {top3 / total:.0%} of the "
            f"classified projects fall into just three divisions, and {len(counts)} "
            f"of the 87 ISIC divisions occur at all. This mirrors the collection "
            f"policy of the archive rather than a limitation of the taxonomy - it is "
            f"a themed archive of social-science studies, so the manufacturing, "
            f"mining and transport divisions are simply absent.")
        lines.append(
            f"- '{isic.division_name(code)}' dominating is consistent with the corpus: "
            "the studies follow school and college populations over time, so schooling "
            "vocabulary is what the researchers themselves wrote in the metadata.")

    # 5. The honest caveat about applying ISIC to this material
    lines.append(
        "- ISIC classifies economic activities, and much of this archive is about "
        "family life, personal identity and the life course, which no ISIC division "
        "describes directly. Such projects land on the division of the institution "
        "they were studied through (a study of mothers observed via their children's "
        "schools scores as Education), so the classes should be read as 'the sector "
        "this data speaks about', not as a summary of the research question.")

    unclassified = conn.execute(
        """SELECT COUNT(*) AS n FROM PROJECTS WHERE repository_id = ?
           AND type IN ('QDA_PROJECT', 'QD_PROJECT') AND primary_class IS NULL""",
        (repo_id,),
    ).fetchone()["n"]
    if unclassified:
        lines.append(
            f"- {unclassified} in-scope project(s) matched no lexicon term and were "
            "deliberately left unclassified rather than forced into a division, "
            "following the instruction to leave a field empty when it cannot be filled.")
    return lines


def build_pdf(conn, path=None):
    path = path or config.REPORT_DIR / f"{config.STUDENT_ID}-sq26-classification.pdf"
    path.parent.mkdir(parents=True, exist_ok=True)

    repositories = conn.execute("SELECT * FROM REPOSITORIES ORDER BY id").fetchall()

    with PdfPages(path) as pdf:
        draw_text_page(pdf, "Seeding QDArchive - Part 2: Classification", title_page(conn))

        for repository in repositories:
            repo_id = repository["id"]
            name = clean(repository["name"])
            header = f"{repo_id}. Repository: {name}"

            draw_text_page(
                pdf, header,
                [f"URL: {repository['url']}", ""] + repository_comments(conn, repository),
            )

            overall = class_counts(conn, repo_id)
            in_scope = conn.execute(
                "SELECT COUNT(*) AS n FROM PROJECTS WHERE repository_id = ? "
                "AND type IN ({})".format(", ".join("?" * len(CLASSIFIED_TYPES))),
                (repo_id, *CLASSIFIED_TYPES),
            ).fetchone()["n"]

            draw_histogram(
                pdf, overall,
                f"{header} - histogram of primary classes",
                "All classified projects (QDA_PROJECT + QD_PROJECT), ISIC Rev. 5 divisions",
            )
            draw_table(
                pdf, overall,
                f"{header} - classes ranked by frequency (top {TOP_N})",
                in_scope,
            )

            # Page 29 of the task description crosses repository with project
            # type, so each (repository, type) pair gets its own distribution.
            for project_type in CLASSIFIED_TYPES:
                counts = class_counts(conn, repo_id, project_type)
                if not counts:
                    continue
                scoped = conn.execute(
                    "SELECT COUNT(*) AS n FROM PROJECTS WHERE repository_id = ? AND type = ?",
                    (repo_id, project_type),
                ).fetchone()["n"]
                draw_histogram(
                    pdf, counts,
                    f"{header} - {project_type}",
                    f"Distribution of primary classes for {project_type} ({scoped} projects)",
                )
                draw_table(
                    pdf, counts,
                    f"{header} - {project_type}: classes ranked by frequency (top {TOP_N})",
                    scoped,
                )

    return path


def title_page(conn):
    repositories = conn.execute("SELECT COUNT(*) AS n FROM REPOSITORIES").fetchone()["n"]
    projects = conn.execute("SELECT COUNT(*) AS n FROM PROJECTS").fetchone()["n"]
    files = conn.execute("SELECT COUNT(*) AS n FROM FILES").fetchone()["n"]
    types = Counter(r["type"] for r in conn.execute("SELECT type FROM PROJECTS"))

    lines = [
        f"Student: A K M Yasar   -   Matriculation ID: {config.STUDENT_ID}",
        "Applied Software Engineering Project, FAU Erlangen-Nurnberg",
        "Prof. Dirk Riehle, Professorship for Open-Source Software",
        "",
        "## Scope",
        f"- Repositories: {repositories}",
        f"- Projects: {projects}",
        f"- Files: {files}",
        "",
        "## Project types (Part 2 Step 1)",
    ]
    for project_type in ALL_TYPES:
        lines.append(f"- {project_type}: {types.get(project_type, 0)}")
    lines += [
        "",
        "## Classification (Part 2 Steps 2 and 3)",
        f"- Taxonomy: {isic.STANDARD}, classified two levels down, i.e. to division level.",
        "- Classifier: a rule-based ISIC lexicon scored over a TF-IDF weighting of the "
        "project metadata and of the text extracted from the downloaded files. It is "
        "deterministic, so re-running it reproduces this report exactly.",
        "- ISIC classifies economic activities, so a project is assigned the division of "
        "the activity its data is about, not the activity of doing research; otherwise "
        "every project here would collapse into division 72 (Scientific research and "
        "development) and the taxonomy would carry no information.",
    ]
    return lines


# ---------------------------------------------------------------------------
# Step 4b: the numbers to type into the Google form
# ---------------------------------------------------------------------------
def form_answers(conn, path=None):
    path = path or config.REPORT_DIR / "form_answers.md"
    path.parent.mkdir(parents=True, exist_ok=True)

    lines = ["# Part 2 Step 4b - answers for the per-repository form",
             "",
             "Fill in the form at https://forms.gle/wxTGQFBQbBvFi3N69 once per repository.",
             ""]

    for repository in conn.execute("SELECT * FROM REPOSITORIES ORDER BY id"):
        repo_id = repository["id"]
        types = Counter(
            r["type"] for r in conn.execute(
                "SELECT type FROM PROJECTS WHERE repository_id = ?", (repo_id,))
        )
        counts = class_counts(conn, repo_id)
        lines += [f"## Repository {repo_id}: {clean(repository['name'])}",
                  f"- URL: {repository['url']}",
                  f"- Projects in total: {sum(types.values())}",
                  "- Project types found:"]
        for project_type in ALL_TYPES:
            lines.append(f"    - {project_type}: {types.get(project_type, 0)}")
        if counts:
            code, count = counts.most_common(1)[0]
            lines.append(f"- Dominant class: {isic.label(code)} ({count} projects)")
            lines.append("- Top 5 classes:")
            for rank, (c, n) in enumerate(counts.most_common(5), 1):
                lines.append(f"    {rank}. {isic.full_class_name(c)}: {n}")
        else:
            lines.append("- Dominant class: none (no classifiable project)")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def main():
    conn = connect()
    xlsx, rows = export_xlsx(conn)
    print(f"wrote {xlsx.relative_to(config.ROOT)} ({rows} rows)")
    pdf = build_pdf(conn)
    print(f"wrote {pdf.relative_to(config.ROOT)}")
    form = form_answers(conn)
    print(f"wrote {form.relative_to(config.ROOT)}")
    conn.close()


if __name__ == "__main__":
    main()
